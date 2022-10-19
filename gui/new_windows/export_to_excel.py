from re import T
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox

from database.database import Database

db = Database("sbu_projects.db")

# === Global Functions ===
selected_id = 0
rows_length = 0

def fn_window_export_to_excel(master):
    # === WINDWOS ===
    window_export_to_excel = tk.Toplevel(master)
    window_export_to_excel.title("Export Database BOM Proyek")
    window_export_to_excel.geometry("695x380")
    window_export_to_excel.resizable(False, False)
    window_export_to_excel.iconbitmap("./favicon.ico")

    # === VARIABLES DECLARATION ===
    EXPORT_FORMAT_OPTIONS = [
        ".xlsx",
        ".xlsx",
        ".csv"
    ]
    format_option_selection = tk.StringVar()
    format_option_selection.set(EXPORT_FORMAT_OPTIONS[0])

    # === FUNCTIONS ===
    def select_item(event):
        try:
            selected = treeview_projects.selection()[0]
            values = treeview_projects.item(selected, 'values')

            # assign selected proejct to variables
            global selected_id
            selected_id = values[0]
            selected_nama_proyek = values[1]
            selected_tahun = values[2]
            selected_kapasitas = values[3]
            selected_customer = values[4]
            selected_jumlah_unit = values[5]
            # selected_jumlah_material = values[6]

            selected_nama_proyek_cropped = selected_nama_proyek
            selected_nama_proyek_cropped = cut_string_10(selected_nama_proyek_cropped)

            entry_selected_project.delete(0, tk.END)
            entry_selected_project.insert(tk.END, selected_nama_proyek)

            # label detail reconfiguration
            label_detail_project_name_selected.configure(text=selected_nama_proyek_cropped)
            label_detail_project_tahun_selected.configure(text=selected_tahun)
            label_detail_project_kapasitas_selected.configure(text=selected_kapasitas)
            label_detail_project_customer_selected.configure(text=selected_customer)
            label_detail_project_jumlah_selected.configure(text=selected_jumlah_unit)
        except:
            pass

    def clear_treeview_project():
        for record in treeview_projects.get_children():
            treeview_projects.delete(record)

    def populate_treeview_project():
        clear_treeview_project()
        count = 0
        for row in db.fetch():
            treeview_projects.insert(parent="", index=count, iid=count, values=row)
            count += 1

    def search_project_to_treeview(param_year, param_name, param_id):
        clear_treeview_project()
        count = 0
        for row in db.search_by_year(param_year, param_name, param_id):
            treeview_projects.insert(parent="", index=count, iid=count, values=row)
            count += 1
    
    def fn_get_export_path():
        global export_file_path
        try:
            filename = filedialog.asksaveasfilename(initialdir="/", title="Ekspor sebagai .xlsx", filetypes=[("Excel file", ".xlsx .xls")])
            entry_path_export.delete(0, tk.END)
            entry_path_export.insert(tk.END, filename)
        except:
            pass

    def cut_string_10(string):
        temp_list = list(string)
        del temp_list[7:(len(temp_list)-10)]
        temp_list.append("...")
        result = "".join(temp_list)
        return result

    def fn_query_all_project_materials(param_project_id):
        try:
            # Pandas dataframe set up
            a = db.fetch_all_material_tables(param_project_id)
            global rows_length
            rows_length = len(a)
            df = pd.DataFrame(a)
            path = entry_path_export.get()
            # Export to excel
            df.to_excel(path, sheet_name="Bill of Materials", startrow=6, startcol=1, header=False, index=False)            
            # Excel formatting
            format_excel_bom(path)
            # Show message info
            messagebox.showinfo("Info", "Berhasil mengekspor")
        except PermissionError:
            messagebox.showerror("Error", "Akses ditolak, silahkan tutup jendela excel")

    # === Excel Formatting Functions
    def format_excel_bom(path):
        workbook = load_workbook(path)
        sheet = workbook["Bill of Materials"]
        # sheet = workbook.active

        thin = Side(border_style="thin", color="00000000")
        double = Side(border_style="medium", color="00000000")

        # content styling
        global rows_length
        for rows in sheet.iter_cols(min_row=7, max_row=int(rows_length), min_col=2, max_col=18):
            print(rows)
            for cell in rows:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, right=thin, bottom=thin, left=thin)
                cell.font = Font(b=False, size=8, name="Times New Roman")
                cell.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")

        # for cell in sheet["7:"]:
        #     cell.alignment = Alignment(horizontal="center", vertical="center")
        #     cell.border = Border(top=thin, right=thin, bottom=thin, left=thin)
        #     cell.font = Font(b=False, size=8, name="Times New Roman")
        #     cell.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")

        sheet.merge_cells("B5:H5")
        sheet.merge_cells("I5:L5")
        sheet.merge_cells("M5:Q5")
        sheet.merge_cells("R5:R6")

        cell_merged_heading_1 = sheet["B5:H5"]
        cell_merged_heading_2 = sheet["I5:L5"]
        cell_merged_heading_3 = sheet["M5:Q5"]
        cell_merged_heading_4 = sheet["R5:R6"]
        for row in cell_merged_heading_1[0]:
            row.border = Border(top=double, right=double, bottom=double, left=double)
        for row in cell_merged_heading_2[0]:
            row.border = Border(top=double, right=double, bottom=double, left=double)
        for row in cell_merged_heading_3[0]:
            row.border = Border(top=double, right=double, bottom=double, left=double)
        for row in cell_merged_heading_4[0]:
            row.border = Border(top=double, right=double, bottom=double, left=double)
        for row in cell_merged_heading_4[1]:
            row.border = Border(top=double, right=double, bottom=double, left=double)

        # header project
        sheet.merge_cells("B4:R4")
        cell_project_header = sheet["B4"]
        cell_project_header.value = str(entry_selected_project.get()).upper()
        cell_project_header.font = Font(name="Times New Roman", b=True, size=24)
        cell_project_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_project_header.border = Border(top=double, right=double, bottom=double, left=double)
        # header selection
        cell_bom_header = sheet["B5"]
        cell_spp_header = sheet["I5"]
        cell_po_header = sheet["M5"]        
        # header values
        cell_bom_header.value = "BOM"
        cell_spp_header.value = "PR/SPP"
        cell_po_header.value = "PO"
        # header styling
        cell_bom_header.font = Font(name="Times New Roman", b=True, size=12)
        cell_spp_header.font = Font(name="Times New Roman", b=True, size=12)
        cell_po_header.font = Font(name="Times New Roman", b=True, size=12)
        # header alignment
        cell_bom_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_spp_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_header.alignment = Alignment(horizontal="center", vertical="center")
        # header border
        cell_merged_title = sheet["B4:R4"]
        for row in cell_merged_title[0]:
            row.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_header.border = Border(top=double, right=double, bottom=double, left=double)
        cell_spp_header.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_header.border = Border(top=double, right=double, bottom=double, left=double)

        # header selection
        cell_bom_no = sheet["B6"]
        cell_bom_rev = sheet["C6"]
        cell_bom_kode = sheet["D6"]
        cell_bom_deskripsi = sheet["E6"]
        cell_bom_spesifikasi = sheet["F6"]
        cell_bom_kuantitas = sheet["G6"]
        cell_bom_satuan = sheet["H6"]
        cell_spp_nomor = sheet["I6"]
        cell_spp_kuantitas = sheet["J6"]
        cell_spp_satuan = sheet["K6"]
        cell_spp_status = sheet["L6"]
        cell_po_nomor = sheet["M6"]
        cell_po_kuantitas = sheet["N6"]
        cell_po_satuan = sheet["O6"]
        cell_po_kode = sheet["P6"]
        cell_po_tanggal = sheet["Q6"]
        cell_stok = sheet["R5"]
        # header values
        cell_bom_no.value = "No."
        cell_bom_rev.value = "Rev"
        cell_bom_kode.value = "Kode Material"
        cell_bom_deskripsi.value = "Deskripsi Material"
        cell_bom_spesifikasi.value = "Spesifikasi"
        cell_bom_kuantitas.value = "Quantity"
        cell_bom_satuan.value = "Satuan"
        cell_spp_nomor.value = "Nomor"
        cell_spp_kuantitas.value = "Quantity"
        cell_spp_satuan.value = "Satuan"
        cell_spp_status.value = "Status"
        cell_po_nomor.value = "Nomor"
        cell_po_kuantitas.value = "Quantity"
        cell_po_satuan.value = "Satuan"
        cell_po_kode.value = "Kode"
        cell_po_tanggal.value = "Tanggal Kedatangan"
        cell_stok.value = "Stok Material/Barang"
        # header font styling
        cell_bom_no.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_rev.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_kode.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_deskripsi.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_spesifikasi.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_kuantitas.font = Font(b=True, size=8, name="Times New Roman")
        cell_bom_satuan.font = Font(b=True, size=8, name="Times New Roman")
        cell_spp_nomor.font = Font(b=True, size=8, name="Times New Roman")
        cell_spp_kuantitas.font = Font(b=True, size=8, name="Times New Roman")
        cell_spp_satuan.font = Font(b=True, size=8, name="Times New Roman")
        cell_spp_status.font = Font(b=True, size=8, name="Times New Roman")
        cell_po_nomor.font = Font(b=True, size=8, name="Times New Roman")
        cell_po_kuantitas.font = Font(b=True, size=8, name="Times New Roman")
        cell_po_satuan.font = Font(b=True, size=8, name="Times New Roman")
        cell_po_kode.font = Font(b=True, size=8, name="Times New Roman")
        cell_po_tanggal.font = Font(b=True, size=8, name="Times New Roman")
        cell_stok.font = Font(b=True, size=8, name="Times New Roman")
        # header alignment
        cell_bom_no.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_rev.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_kode.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_deskripsi.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_spesifikasi.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_kuantitas.alignment = Alignment(horizontal="center", vertical="center")
        cell_bom_satuan.alignment = Alignment(horizontal="center", vertical="center")
        cell_spp_nomor.alignment = Alignment(horizontal="center", vertical="center")
        cell_spp_kuantitas.alignment = Alignment(horizontal="center", vertical="center")
        cell_spp_satuan.alignment = Alignment(horizontal="center", vertical="center")
        cell_spp_status.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_nomor.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_kuantitas.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_satuan.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_kode.alignment = Alignment(horizontal="center", vertical="center")
        cell_po_tanggal.alignment = Alignment(horizontal="center", vertical="center")
        cell_stok.alignment = Alignment(horizontal="center", vertical="center")
        # header border
        cell_bom_no.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_rev.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_kode.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_deskripsi.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_spesifikasi.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_kuantitas.border = Border(top=double, right=double, bottom=double, left=double)
        cell_bom_satuan.border = Border(top=double, right=double, bottom=double, left=double)
        cell_spp_nomor.border = Border(top=double, right=double, bottom=double, left=double)
        cell_spp_kuantitas.border = Border(top=double, right=double, bottom=double, left=double)
        cell_spp_satuan.border = Border(top=double, right=double, bottom=double, left=double)
        cell_spp_status.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_nomor.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_kuantitas.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_satuan.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_kode.border = Border(top=double, right=double, bottom=double, left=double)
        cell_po_tanggal.border = Border(top=double, right=double, bottom=double, left=double)
        cell_stok.border = Border(top=double, right=double, bottom=double, left=double)

        workbook.save(path)

    # === WIDGETS ===
    # Treeview List Proyek
    treeview_projects = ttk.Treeview(window_export_to_excel, height=6)

    treeview_projects["column"] = ("ID", "Nama Proyek", "Tahun", "Kapasitas", "Customer", "Jumlah Unit")
    treeview_projects.column("#0", anchor=tk.CENTER, width=0)
    treeview_projects.column("ID", anchor=tk.CENTER, width=45)        
    treeview_projects.column("Nama Proyek", anchor=tk.CENTER, width=125)
    treeview_projects.column("Tahun", anchor=tk.CENTER, width=125)
    treeview_projects.column("Kapasitas", anchor=tk.CENTER, width=125)
    treeview_projects.column("Customer", anchor=tk.CENTER, width=125)
    treeview_projects.column("Jumlah Unit", anchor=tk.CENTER, width=125)
    # Heading declaration
    treeview_projects.heading("#0", text="", anchor=tk.CENTER)
    treeview_projects.heading("ID", text="ID", anchor=tk.CENTER)
    treeview_projects.heading("Nama Proyek", text="Nama Proyek", anchor=tk.CENTER)
    treeview_projects.heading("Tahun", text="Tahun", anchor=tk.CENTER)
    treeview_projects.heading("Kapasitas", text="Kapasitas", anchor=tk.CENTER)
    treeview_projects.heading("Customer", text="Customer", anchor=tk.CENTER)
    treeview_projects.heading("Jumlah Unit", text="Jumlah Unit", anchor=tk.CENTER)
    # Bind functions
    treeview_projects.bind("<<TreeviewSelect>>", select_item)
    treeview_projects.grid(row=0, column=0, columnspan=2, padx=11, pady=6)
    # Framegroup -> Cari Proyek
    framegroup_search_project = tk.LabelFrame(window_export_to_excel, text="Cari proyek")
    framegroup_search_project.grid(row=1, column=0, padx=11, sticky=tk.N,)
    # nama proyek
    label_nama = tk.Label(framegroup_search_project, text="Nama Proyek: ")
    label_nama.grid(row=0, column=0, sticky=tk.W)
    entry_nama = tk.Entry(framegroup_search_project, width=40)
    entry_nama.grid(row=0, column=1, sticky=tk.W)
    # proyek terpilih
    label_selected_project = tk.Label(framegroup_search_project, text="Proyek Terpilih: ")
    label_selected_project.grid(row=1, column=0, sticky=tk.W)
    entry_selected_project = tk.Entry(framegroup_search_project, width=40)
    entry_selected_project.grid(row=1, column=1, columnspan=2, sticky=tk.W)
    # tombol -> button group
    button_search = ttk.Button(framegroup_search_project, text="Cari", width=9, command=lambda: search_project_to_treeview("SEMUA", entry_nama.get(), 0))
    button_search.grid(row=0, column=2, padx=6, sticky=tk.E)
    # button -> search all
    button_search_all = ttk.Button(framegroup_search_project, text="All", width=9, command=lambda: search_project_to_treeview("SEMUA", "", 0))
    button_search_all.grid(row=1, column=2, padx=6, sticky=tk.E, pady=6)
    # Framegroup -> Selected project detalils
    framegroup_project_info = tk.LabelFrame(window_export_to_excel, text="Rincian", font=("Verdana bold", 9), width=30)
    framegroup_project_info.grid(row=1, column=1, sticky=tk.W, rowspan=2)
    # Project details
    # nama proyek -> selected
    label_detail_project_name_selected = tk.Label(framegroup_project_info, text="<Silahkan pilih proyek>", font=("Arial", 9), width=27)
    label_detail_project_name_selected.grid(row=0, column=0, columnspan=2, padx=3)
    # tahun
    label_detail_project_tahun = tk.Label(framegroup_project_info, text="Tahun: ", font=("Arial", 9))
    label_detail_project_tahun.grid(row=1, column=0, sticky=tk.E)
    # tahun -> selected
    label_detail_project_tahun_selected = tk.Label(framegroup_project_info, text="-", font=("Arial", 9), width=10)
    label_detail_project_tahun_selected.grid(row=1, column=1, sticky=tk.E, padx=3)
    # kapasitas
    label_detail_kapasitas = tk.Label(framegroup_project_info, text="Kapasitas: ", font=("Arial", 9))
    label_detail_kapasitas.grid(row=2, column=0, sticky=tk.E)
    # kapasitas -> selected
    label_detail_project_kapasitas_selected = tk.Label(framegroup_project_info, text="-", font=("Arial", 9), width=10)
    label_detail_project_kapasitas_selected.grid(row=2, column=1, sticky=tk.E, padx=3)
    # customer
    label_detail_project_customer = tk.Label(framegroup_project_info, text="Customer: ", font=("Arial", 9), width=10)
    label_detail_project_customer.grid(row=3, column=0, sticky=tk.E)
    # custoer -> selected
    label_detail_project_customer_selected = tk.Label(framegroup_project_info, text="-", font=("Arial", 9), width=10)
    label_detail_project_customer_selected.grid(row=3, column=1, sticky=tk.E, padx=3)
    # jumlah unit
    label_detail_project_jumlah = tk.Label(framegroup_project_info, text="ΣUnit: ", font=("Arial", 9), width=10)
    label_detail_project_jumlah.grid(row=4, column=0, sticky=tk.E)
    # jumlah unit -> selected
    label_detail_project_jumlah_selected = tk.Label(framegroup_project_info, text="-", font=("Arial", 9), width=10)
    label_detail_project_jumlah_selected.grid(row=4, column=1, sticky=tk.E, padx=3)
    # jumlah material
    label_detail_jumlah_bom = tk.Label(framegroup_project_info, text="ΣMaterial: ", font=("Arial", 9), width=10)
    label_detail_jumlah_bom.grid(row=5, column=0, sticky=tk.E)
    # jumlah material -> selected
    label_detail_jumlah_bom_selected = tk.Label(framegroup_project_info, text="-", font=("Arial", 9), width=10)
    label_detail_jumlah_bom_selected.grid(row=5, column=1, sticky=tk.E, padx=3)
    # datasheet link
    label_detail_project_datasheet = tk.Label(framegroup_project_info, text="Data Sheet", font=("Arial", 9), fg="blue")
    label_detail_project_datasheet.grid(row=6, column=0, columnspan=2, pady=8)
    # labelframe export setting
    labelframe_export_setting = tk.LabelFrame(window_export_to_excel, text="Pengaturan")
    labelframe_export_setting.grid(row=2, column=0, sticky=tk.N)
    # path export
    label_path_export = tk.Label(labelframe_export_setting, text="Path :")
    label_path_export.grid(row=0, column=0, sticky=tk.W, padx=5)
    entry_path_export = tk.Entry(labelframe_export_setting, width=34)
    entry_path_export.grid(row=0, column=1, sticky=tk.E, padx=6)
    # path export button
    button_path_export = ttk.Button(labelframe_export_setting, text="Pilih", command=fn_get_export_path)
    button_path_export.grid(row=0, column=2, padx=6)
    # format ekspor
    label_export_format = tk.Label(labelframe_export_setting, text="Format ekspor: ", background="cyan")
    label_export_format.grid(row=1, column=0, sticky=tk.W, padx=5)
    # format ekspor -> menu
    optionmenu_export_format = ttk.OptionMenu(labelframe_export_setting, format_option_selection, *EXPORT_FORMAT_OPTIONS)
    optionmenu_export_format.grid(row=1, column=1, sticky=tk.E)
    # simpan log ekspor
    # Button export
    button_export_excel = ttk.Button(window_export_to_excel, text="Ekspor", width=72, command=lambda: fn_query_all_project_materials(selected_id))
    button_export_excel.grid(row=3, column=0, columnspan=2, pady=6)
    # button_export_excel.grid()

    # === INITIAL WINDOWS STARTUP FUNCTION CALL ===
    populate_treeview_project()